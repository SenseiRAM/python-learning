import credentials # private login kept in separate py file
import openpyxl
import os
import time
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime

os.chdir('C:\\Users\\richa\\Documents\\Python Scripts')
wb = openpyxl.load_workbook('eventsheet.xlsx')
ws = wb.active

class Event():
    """This object contains Eventbrite data pulled from our spreadsheet"""
    ticket_end_time = '5:00pm'

    def __init__(self, title, venue, start_date, end_date, start_time, end_time, short_sum, desc, seats, rsvp_deadline):

        self.title = title
        self.venue = venue
        self.start_date = start_date
        self.start_time = start_time
        self.end_date = end_date
        self.end_time = end_time
        self.short_sum = short_sum
        self.desc = desc
        self.seats = seats
        self.rsvp_deadline = rsvp_deadline

#event = Event(title=ws['A1'].value, venue=ws['B1'].value, start_date=ws['C1'].value, end_date=ws['D1'].value, start_time=ws['E1'].value, end_time=ws['F1'].value, short_sum=ws['G1'].value, desc=ws['H1'].value, seats=ws['I1'].value, rsvp_deadline=ws['J1'].value)
event = Event(ws['A1'].value, ws['B1'].value, ws['C1'].value, ws['D1'].value, ws['E1'].value, ws['F1'].value, ws['G1'].value, ws['H1'].value, ws['I1'].value, ws['J1'].value)

driver = webdriver.Chrome('C:\\Users\\richa\\Documents\\Python Scripts\\chromedriver')

driver.get("https://www.eventbrite.com/login/")
user_email = driver.find_element_by_id("email")
user_email.send_keys(credentials.eb_name, Keys.ENTER)

pw = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "password"))
        )
pw.send_keys(credentials.eb_pw, Keys.ENTER)
time.sleep(4)
driver.get("https://www.eventbrite.com/create/")
modal_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//button[@data-automation="modal-close-button"]'))
        )
#modal_button = driver.find_element_by_xpath('//button[@data-automation="modal-close-button"]')
modal_button.click()
event_title = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "event-basicInfo-title"))
        )
event_title.send_keys(event.title)
venue = driver.find_element_by_id("event-locationautocomplete-location")
venue.send_keys(event.venue)
time.sleep(2)
# Code executes too fast if we keep this on the same line
venue.send_keys(Keys.ARROW_DOWN, Keys.ENTER)

start_date_field = driver.find_element_by_id("event-startDate")
start_date_field.send_keys(Keys.CONTROL + "a")
start_date_field.send_keys(event.start_date.strftime("%m/%d/%Y"))
start_time_field = driver.find_element_by_id("event-startTime")
start_time_field.send_keys(Keys.CONTROL + "a")
start_time_field.send_keys(str(event.start_time))

end_date_field = driver.find_element_by_id("event-endDate")
end_date_field.send_keys(Keys.CONTROL + "a")
end_date_field.send_keys(event.end_date.strftime("%m/%d/%Y"))
end_time_field = driver.find_element_by_id("event-endTime")
end_time_field.send_keys(Keys.CONTROL + "a")
end_time_field.send_keys(str(event.end_time))

save_button = driver.find_element_by_xpath('//button[@data-automation="coyote-action-save"]')
save_button.click()



short_sum_fld = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "event-design-description"))
        )
short_sum_fld.send_keys(event.short_sum)

desc_fld = driver.find_element_by_class_name("eds-richtexteditor__input")
desc_fld.send_keys(event.desc)

save_button.click()

free_ticket_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//label[@for="segmented-ticketType-1"]'))
        )
free_ticket_button.click()

ticket_name = driver.find_element_by_id('ticket-name')
ticket_name.send_keys(Keys.CONTROL + 'a')
ticket_name.send_keys('Register')

ticket_number = driver.find_element_by_id('ticket-quantity')
ticket_number.send_keys(event.seats)

ticket_end_date = driver.find_element_by_id('ticket-sales-end-date')
ticket_end_date.send_keys(Keys.CONTROL + 'a')
ticket_end_date.send_keys(event.rsvp_deadline.strftime("%m/%d/%Y"))

ticket_end_time_fld = driver.find_element_by_id('ticket-sales-end-time')
ticket_end_time_fld.send_keys(Keys.CONTROL + 'a')
ticket_end_time_fld.send_keys(event.ticket_end_time)

ticket_save_button = driver.find_element_by_xpath('//button[@data-automation="coyote-ticket-form-action-save"]')
ticket_save_button.click()
time.sleep(5)
privacy_menu = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//div[text()="Privacy Settings"]'))
        )
privacy_menu.click()
driver.find_element_by_xpath('//div[text()="Privacy"]').click()

private_dropdown = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//select[@data-automation="coyote-privacy-dropdown"]'))
        )
private_dropdown.click()
private_dropdown.send_keys(Keys.ARROW_DOWN, Keys.ENTER)

privacy_save_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//button[@data-automation="coyote-action-save"]'))
        )
privacy_save_button.click()
time.sleep(2)

publish_menu = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//div[text()="Publish Event"]'))
        )
publish_menu.click()
